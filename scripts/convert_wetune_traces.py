#!/usr/bin/env python3
"""
将 WeTune RewriteQuery 生成的 TSV trace 文件转换为 Excel 格式。

每条查询（stmtId）生成一个 .xlsx 文件，格式与 diaspora_rewrite_paths.xlsx 一致：
  - 每个 sheet = 一条改写路径（Path_0, Path_1, ...）
  - 每个 sheet 的列：seq | ruleId | ruleDesc（规则模板字符串）

TSV 输入格式（tab 分隔）：
  appName  stmtId  rewriteIndex  optimizedSql  trace(逗号分隔规则行号)

用法:
  python3 scripts/convert_wetune_traces.py \
      --tsv /tmp/wetune_traces/run.../1_query.tsv \
      --rules /mnt/mydata/yyk/AutoRewriter/test/src/main/resources/wetune_rules.txt \
      --output-root /mnt/mydata/yyk/AutoRewriter/test/src/main/resources/schema \
      [--apps diaspora discourse]
"""

import argparse
import sys
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def load_rules(rules_path: Path) -> dict:
    """加载 rules.txt，返回 {行号(1-based): 规则模板字符串} 的映射。
    将规则中的 k0,k1,... 占位符替换为 a0,a1,... 以兼容 AutoRewriter 解析器。
    """
    import re
    rules = {}
    with rules_path.open(encoding="utf-8") as f:
        for i, line in enumerate(f, 1):
            line = line.strip()
            if line:
                # 将 k0,k1,... 替换为 a0,a1,...（与 RewritePathE2ETest.replaceKPlaceholders 一致）
                line = re.sub(r'\bk(\d+)', lambda m: 'a' + m.group(1), line)
                rules[i] = line
    return rules


def parse_trace(trace_str: str) -> list:
    """解析 trace 列，返回规则行号列表（过滤掉 -1 等预处理规则）"""
    if not trace_str or not trace_str.strip():
        return []
    result = []
    for x in trace_str.strip().split(","):
        x = x.strip()
        try:
            n = int(x)
            if n > 0:
                result.append(n)
        except ValueError:
            pass
    return result


def write_excel(output_path: Path, stmt_id: int, paths: list, rules_map: dict):
    """
    写出一个 Excel 文件，每条路径一个 sheet。

    paths: list of {"pathIndex": int, "ruleLines": [int, ...], "optimizedSql": str}
    """
    wb = openpyxl.Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    for path in sorted(paths, key=lambda p: p["pathIndex"]):
        sheet_name = f"Path_{path['pathIndex']}"
        ws = wb.create_sheet(title=sheet_name)

        # 写表头
        headers = ["seq", "ruleId", "ruleDesc"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 写 Original 行（第一行数据）
        ws.cell(row=2, column=1, value="Original")
        ws.cell(row=2, column=2, value="")
        ws.cell(row=2, column=3, value="")

        # 写每条规则
        rule_lines = path["ruleLines"]
        for seq_idx, rule_line in enumerate(rule_lines, 1):
            row = seq_idx + 2  # 跳过表头行和 Original 行
            rule_template = rules_map.get(rule_line, f"<unknown rule line {rule_line}>")
            ws.cell(row=row, column=1, value=str(seq_idx))
            ws.cell(row=row, column=2, value=str(rule_line))
            ws.cell(row=row, column=3, value=rule_template)

        # 调整列宽
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 120

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Convert WeTune TSV traces to Excel (per-query .xlsx)")
    parser.add_argument("--tsv", required=True, type=Path, help="Input TSV from WeTune RewriteQuery")
    parser.add_argument("--rules", required=True, type=Path, help="WeTune rules.txt file")
    parser.add_argument("--output-root", required=True, type=Path,
                        help="Root schema dir (e.g. test/src/main/resources/schema)")
    parser.add_argument("--apps", nargs="+", default=None, help="Only convert these apps")
    args = parser.parse_args()

    if not args.tsv.exists():
        print(f"ERROR: TSV not found: {args.tsv}", file=sys.stderr)
        sys.exit(1)
    if not args.rules.exists():
        print(f"ERROR: rules.txt not found: {args.rules}", file=sys.stderr)
        sys.exit(1)

    print(f"Loading rules from {args.rules}...")
    rules_map = load_rules(args.rules)
    print(f"  Loaded {len(rules_map)} rules")

    # 按 (app, stmtId) 分组
    stmt_paths = defaultdict(list)

    print(f"Reading TSV from {args.tsv}...")
    with args.tsv.open(encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            parts = line.split("\t")
            if len(parts) < 5:
                continue
            app_name, stmt_id_str, rewrite_idx_str, opt_sql, trace_str = (
                parts[0], parts[1], parts[2], parts[3], parts[4]
            )
            if args.apps and app_name not in args.apps:
                continue
            try:
                stmt_id = int(stmt_id_str)
                rewrite_idx = int(rewrite_idx_str)
            except ValueError:
                continue

            rule_lines = parse_trace(trace_str)
            if not rule_lines:
                continue  # 跳过无规则触发的路径

            stmt_paths[(app_name, stmt_id)].append({
                "pathIndex": rewrite_idx,
                "ruleLines": rule_lines,
                "optimizedSql": opt_sql,
            })

    print(f"Found {len(stmt_paths)} (app, stmtId) pairs with traces")

    # 写出 Excel 文件
    written = 0
    app_counts = defaultdict(int)

    for (app_name, stmt_id), paths in sorted(stmt_paths.items()):
        wetune_app = f"wetune_{app_name}"
        traces_dir = args.output_root / wetune_app / "traces"
        output_path = traces_dir / f"{stmt_id}.xlsx"

        write_excel(output_path, stmt_id, paths, rules_map)
        written += 1
        app_counts[app_name] += 1

        if written % 100 == 0:
            print(f"  Written {written} files...")

    print(f"\nDone. Written {written} Excel files.")
    for app, count in sorted(app_counts.items()):
        print(f"  {app}: {count} queries with traces")


if __name__ == "__main__":
    main()
